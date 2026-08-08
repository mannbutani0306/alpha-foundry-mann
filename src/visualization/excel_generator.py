import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_excel_validation_workbook(data_path="data/processed/processed_factors.parquet"):
    print("Generating Excel Validation Workbook (Deliverable 5)...")
    if not os.path.exists(data_path):
        raise FileNotFoundError(f"Data file not found at {data_path}.")

    df = pd.read_parquet(data_path)
    
    # Pick a single recent sample date for clear manual audit verification
    sample_date = sorted(df['date'].unique())[-1]
    sample_df = df[df['date'] == sample_date].copy().head(15).reset_index(drop=True)

    # Select key representative factor columns
    cols_to_keep = ['date', 'Ticker', 'close', 'ret_1m', 'ret_3m', 'vol_20d', 'fwd_ret_21d']
    sample_df = sample_df[[c for c in cols_to_keep if c in sample_df.columns]]

    # 1. Compute Manual Z-Scores for Audit Sheet
    for factor in ['ret_1m', 'ret_3m', 'vol_20d']:
        if factor in sample_df.columns:
            mean_val = sample_df[factor].mean()
            std_val = sample_df[factor].std(ddof=1)
            sample_df[f'{factor}_z_manual'] = (sample_df[factor] - mean_val) / (std_val + 1e-9)

    # 2. Compute Composite Rank and Selection Flag
    z_cols = [c for c in sample_df.columns if c.endswith('_z_manual')]
    if z_cols:
        sample_df['composite_score'] = sample_df[z_cols].mean(axis=1)
        sample_df['rank'] = sample_df['composite_score'].rank(ascending=False, method='min')
        sample_df['shortlist_flag'] = np.where(sample_df['rank'] <= 3, 'SELECTED', 'REJECTED')

    wb = openpyxl.Workbook()
    
    # Sheet 1: Manual Calculation Audit
    ws1 = wb.active
    ws1.title = "Factor_ZScore_Audit"

    # Styling definitions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Write dataframe rows to Sheet 1
    for r in dataframe_to_rows(sample_df, index=False, header=True):
        ws1.append(r)

    # Format Header Row
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Format Data Rows
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if isinstance(cell.value, float):
                cell.number_format = '0.0000'

    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Sheet 2: Summary Metrics & Formulas
    ws2 = wb.create_sheet(title="IC_Summary_Metrics")
    ws2.append(["Metric Name", "Value", "Formula / Description"])
    ws2.append(["Sample Audit Date", str(sample_date), "Date used for single-period cross-sectional audit"])
    ws2.append(["Total Names in Sample", len(sample_df), "Number of stocks in sample universe"])
    ws2.append(["Rank IC (Spearman)", "=CORREL(H2:H16, G2:G16)", "Excel CORREL between Composite Score and Forward Returns"])
    ws2.append(["Top Shortlist Cutoff", 3, "Top-N names flagged for long allocation"])

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border

    # Save directly to the top-level 'reports/' directory shown in your tree
    os.makedirs("reports", exist_ok=True)
    save_path = "reports/validation_workbook.xlsx"
    wb.save(save_path)

    print("--------------------------------------------------")
    print(f"Excel Validation Workbook successfully generated!")
    print(f"Saved workbook to: '{save_path}'")
    print("--------------------------------------------------")

if __name__ == "__main__":
    generate_excel_validation_workbook()